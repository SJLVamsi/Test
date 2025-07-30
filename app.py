import flask
from flask import Flask, render_template, request, redirect, session, make_response, send_file
import sqlite3
import os
import json
import openpyxl
from fpdf import FPDF
import matplotlib.pyplot as plt
from datetime import datetime
import io
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.secret_key = 'secret123'

# Configuration for file uploads
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'pdf'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Create upload directory if it doesn't exist
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def init_db():
    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS users (id INTEGER PRIMARY KEY, username TEXT, password TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS results (
                    id INTEGER PRIMARY KEY, 
                    username TEXT, 
                    score INTEGER, 
                    answers TEXT, 
                    start_time TEXT, 
                    end_time TEXT, 
                    time_per_question TEXT
                 )''')
    c.execute('''CREATE TABLE IF NOT EXISTS admin (id INTEGER PRIMARY KEY, username TEXT, password TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS questions (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    section INTEGER NOT NULL,
                    question TEXT NOT NULL,
                    option_a TEXT NOT NULL,
                    option_b TEXT NOT NULL,
                    option_c TEXT NOT NULL,
                    option_d TEXT NOT NULL,
                    correct_answer TEXT NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    is_active BOOLEAN DEFAULT 1
                 )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS documents (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    title TEXT NOT NULL,
                    description TEXT,
                    filename TEXT NOT NULL,
                    file_path TEXT NOT NULL,
                    file_size INTEGER,
                    uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    is_active BOOLEAN DEFAULT 1
                 )''')
    
    # Check if admin user exists
    c.execute("DELETE FROM admin WHERE username = 'admin'")
    c.execute("SELECT * FROM admin WHERE username = 'vamsi'")
    if c.fetchone() is None:
        c.execute("INSERT INTO admin (username, password) VALUES ('vamsi', '1234')")
    
    # Check if questions table is empty and migrate from JSON if needed
    c.execute("SELECT COUNT(*) FROM questions")
    if c.fetchone()[0] == 0:
        migrate_questions_from_json(c)
    
    conn.commit()
    conn.close()

def migrate_questions_from_json(cursor):
    """Migrate existing questions from JSON to database"""
    try:
        with open('questions.json') as f:
            data = json.load(f)
            questions = data['questions']
            for q in questions:
                cursor.execute('''INSERT INTO questions 
                                 (section, question, option_a, option_b, option_c, option_d, correct_answer) 
                                 VALUES (?, ?, ?, ?, ?, ?, ?)''',
                              (q['section'], q['question'], 
                               q['options']['a'], q['options']['b'], 
                               q['options']['c'], q['options']['d'], 
                               q['answer']))
    except FileNotFoundError:
        # If JSON file doesn't exist, create some sample questions
        sample_questions = [
            (1, "What is the capital of France?", "Berlin", "Madrid", "Paris", "Rome", "c"),
            (1, "What is 2 + 2?", "3", "4", "5", "6", "b"),
            (1, "What is the largest planet in our solar system?", "Earth", "Jupiter", "Mars", "Saturn", "b")
        ]
        for q in sample_questions:
            cursor.execute('''INSERT INTO questions 
                             (section, question, option_a, option_b, option_c, option_d, correct_answer) 
                             VALUES (?, ?, ?, ?, ?, ?, ?)''', q)

def load_questions():
    """Load questions from database instead of JSON"""
    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    c.execute('''SELECT id, section, question, option_a, option_b, option_c, option_d, correct_answer 
                 FROM questions WHERE is_active = 1 ORDER BY section, id''')
    rows = c.fetchall()
    conn.close()
    
    questions = []
    for row in rows:
        q_id, section, question, opt_a, opt_b, opt_c, opt_d, answer = row
        questions.append({
            'id': q_id,
            'section': section,
            'question': question,
            'options': {
                'a': opt_a,
                'b': opt_b,
                'c': opt_c,
                'd': opt_d
            },
            'answer': answer
        })
    return questions

def log_login(username):
    if not os.path.exists('login_details.xlsx'):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Login Details"
        sheet.append(['Username', 'Login Time'])
        workbook.save('login_details.xlsx')
    
    workbook = openpyxl.load_workbook('login_details.xlsx')
    sheet = workbook.active
    sheet.append([username, datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    workbook.save('login_details.xlsx')

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        conn = sqlite3.connect('database.db')
        c = conn.cursor()
        c.execute('INSERT INTO users (username, password) VALUES (?, ?)', (username, password))
        conn.commit()
        conn.close()
        return redirect('/login')
    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        conn = sqlite3.connect('database.db')
        c = conn.cursor()
        c.execute('SELECT * FROM users WHERE username=? AND password=?', (username, password))
        user = c.fetchone()
        conn.close()
        if user:
            session['username'] = username
            log_login(username)
            session['start_time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            return redirect('/test')
    return render_template('login.html')

@app.route('/test', methods=['GET', 'POST'])
def test():
    if 'username' not in session:
        return redirect('/login')
    
    questions = load_questions()
    
    if request.method == 'POST':
        answers = request.form
        end_time = datetime.now()
        start_time_str = session.get('start_time')
        start_time = datetime.strptime(start_time_str, '%Y-%m-%d %H:%M:%S')
        
        time_per_question = {}
        for question in questions:
            time_per_question[f'q{question["id"]}'] = request.form.get(f'q{question["id"]}_time', '0')

        score = 0
        user_answers = {}
        for q in questions:
            user_answer = answers.get(f'q{q["id"]}')
            user_answers[f'q{q["id"]}'] = {
                "question": q['question'],
                "options": q['options'],
                "selected": user_answer,
                "correct_answer": q['answer'],
                "time_taken": time_per_question.get(f'q{q["id"]}', '0')
            }
            if user_answer == q['answer']:
                score += 1

        total_questions = len(questions)
        score_percentage = round((score / total_questions) * 100) if total_questions > 0 else 0

        conn = sqlite3.connect('database.db')
        c = conn.cursor()
        c.execute('''INSERT INTO results (username, score, answers, start_time, end_time, time_per_question) 
                     VALUES (?, ?, ?, ?, ?, ?)''', 
                  (session['username'], score_percentage, json.dumps(user_answers), start_time_str, end_time.strftime('%Y-%m-%d %H:%M:%S'), json.dumps(time_per_question)))
        conn.commit()
        result_id = c.lastrowid
        conn.close()
        
        return redirect(f'/report/{result_id}')

    return render_template('test.html', questions=questions)

@app.route('/report/<int:result_id>')
def report(result_id):
    if not session.get('admin_logged_in'):
        if 'username' in session:
            return "Your test has been submitted. Only an admin can view the detailed report."
        return redirect('/admin/login')

    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    c.execute('SELECT * FROM results WHERE id=?', (result_id,))
    result = c.fetchone()
    conn.close()

    if not result:
        return "Report not found", 404

    username, score, answers_json, start_time, end_time, time_per_question_json = result[1:]
    answers = json.loads(answers_json)
    time_per_question = json.loads(time_per_question_json)

    # Generate PDF
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    pdf.cell(200, 10, txt=f"Test Report for {username}", ln=True, align='C')
    pdf.cell(200, 10, txt=f"Score: {score}/40", ln=True)
    pdf.cell(200, 10, txt=f"Start Time: {start_time}", ln=True)
    pdf.cell(200, 10, txt=f"End Time: {end_time}", ln=True)
    
    total_time = sum(int(t) for t in time_per_question.values())
    pdf.cell(200, 10, txt=f"Total Time Spent: {total_time} seconds", ln=True)

    pdf.ln(10)
    pdf.cell(200, 10, txt="Question Analysis", ln=True)
    
    correct_count = 0
    wrong_count = 0
    unattempted_count = 0

    for q_id, data in answers.items():
        question_num = q_id.replace('q', '')
        pdf.ln(5)
        pdf.multi_cell(0, 5, f"Q{question_num}: {data['question']}")
        
        for opt, text in data['options'].items():
            pdf.cell(0, 5, f"  {opt}) {text}", ln=True)
            
        pdf.cell(0, 5, f"  Selected Answer: {data['selected'] if data['selected'] else 'Unattempted'}", ln=True)
        pdf.cell(0, 5, f"  Correct Answer: {data['correct_answer']}", ln=True)
        pdf.cell(0, 5, f"  Time Taken: {data['time_taken']} seconds", ln=True)
        
        if data['selected'] == data['correct_answer']:
            correct_count += 1
            pdf.cell(0, 5, "  Result: Correct (+1)", ln=True)
        elif not data['selected']:
            unattempted_count += 1
            pdf.cell(0, 5, "  Result: Unattempted (0)", ln=True)
        else:
            wrong_count += 1
            pdf.cell(0, 5, "  Result: Wrong (0)", ln=True)

    # Pie chart for efficiency
    labels = 'Correct', 'Wrong', 'Unattempted'
    sizes = [correct_count, wrong_count, unattempted_count]
    colors = ['#99ff99','#ff9999','#66b3ff']
    
    fig1, ax1 = plt.subplots()
    ax1.pie(sizes, colors=colors, labels=labels, autopct='%1.1f%%', startangle=90)
    ax1.axis('equal')
    
    chart_path = f'efficiency_{result_id}.png'
    plt.savefig(chart_path)
    plt.close()

    pdf.add_page()
    pdf.set_font("Arial", 'B', size=16)
    pdf.cell(200, 10, txt="Efficiency Chart", ln=True, align='C')
    pdf.image(chart_path, x=10, y=30, w=190)
    os.remove(chart_path)

    response = make_response(pdf.output(dest='S').encode('latin-1'))
    response.headers.set('Content-Disposition', 'attachment', filename=f'report_{username}_{result_id}.pdf')
    response.headers.set('Content-Type', 'application/pdf')
    return response

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        conn = sqlite3.connect('database.db')
        c = conn.cursor()
        c.execute('SELECT * FROM admin WHERE username=? AND password=?', (username, password))
        admin = c.fetchone()
        conn.close()
        if admin:
            session['admin_logged_in'] = True
            return redirect('/admin')
    return render_template('admin_login.html')

@app.route('/admin')
def admin():
    if not session.get('admin_logged_in'):
        return redirect('/admin/login')
    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    c.execute('SELECT id, username, score, start_time, end_time FROM results')
    results = c.fetchall()
    c.execute('SELECT COUNT(*) FROM questions WHERE is_active = 1')
    question_count = c.fetchone()[0]
    conn.close()
    return render_template('admin.html', results=results, question_count=question_count)

@app.route('/admin/questions')
def admin_questions():
    if not session.get('admin_logged_in'):
        return redirect('/admin/login')
    
    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    c.execute('''SELECT id, section, question, option_a, option_b, option_c, option_d, correct_answer, created_at 
                 FROM questions WHERE is_active = 1 ORDER BY section, id''')
    questions = c.fetchall()
    conn.close()
    return render_template('admin_questions.html', questions=questions)

@app.route('/admin/questions/add', methods=['GET', 'POST'])
def admin_add_question():
    if not session.get('admin_logged_in'):
        return redirect('/admin/login')
    
    if request.method == 'POST':
        section = request.form['section']
        question = request.form['question']
        option_a = request.form['option_a']
        option_b = request.form['option_b']
        option_c = request.form['option_c']
        option_d = request.form['option_d']
        correct_answer = request.form['correct_answer']
        
        conn = sqlite3.connect('database.db')
        c = conn.cursor()
        c.execute('''INSERT INTO questions 
                     (section, question, option_a, option_b, option_c, option_d, correct_answer) 
                     VALUES (?, ?, ?, ?, ?, ?, ?)''',
                  (section, question, option_a, option_b, option_c, option_d, correct_answer))
        conn.commit()
        conn.close()
        return redirect('/admin/questions')
    
    return render_template('admin_add_question.html')

@app.route('/admin/questions/edit/<int:question_id>', methods=['GET', 'POST'])
def admin_edit_question(question_id):
    if not session.get('admin_logged_in'):
        return redirect('/admin/login')
    
    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    
    if request.method == 'POST':
        section = request.form['section']
        question = request.form['question']
        option_a = request.form['option_a']
        option_b = request.form['option_b']
        option_c = request.form['option_c']
        option_d = request.form['option_d']
        correct_answer = request.form['correct_answer']
        
        c.execute('''UPDATE questions SET 
                     section=?, question=?, option_a=?, option_b=?, option_c=?, option_d=?, correct_answer=?
                     WHERE id=?''',
                  (section, question, option_a, option_b, option_c, option_d, correct_answer, question_id))
        conn.commit()
        conn.close()
        return redirect('/admin/questions')
    
    c.execute('SELECT * FROM questions WHERE id=?', (question_id,))
    question = c.fetchone()
    conn.close()
    
    if not question:
        return "Question not found", 404
    
    return render_template('admin_edit_question.html', question=question)

@app.route('/admin/questions/delete/<int:question_id>')
def admin_delete_question(question_id):
    if not session.get('admin_logged_in'):
        return redirect('/admin/login')
    
    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    c.execute('UPDATE questions SET is_active = 0 WHERE id = ?', (question_id,))
    conn.commit()
    conn.close()
    return redirect('/admin/questions')

@app.route('/admin/documents')
def admin_documents():
    if not session.get('admin_logged_in'):
        return redirect('/admin/login')
    
    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    c.execute('''SELECT id, title, description, filename, file_size, uploaded_at 
                 FROM documents WHERE is_active = 1 ORDER BY uploaded_at DESC''')
    documents = c.fetchall()
    conn.close()
    return render_template('admin_documents.html', documents=documents)

@app.route('/admin/documents/upload', methods=['GET', 'POST'])
def admin_upload_document():
    if not session.get('admin_logged_in'):
        return redirect('/admin/login')
    
    if request.method == 'POST':
        title = request.form['title']
        description = request.form['description']
        
        # Check if file was uploaded
        if 'file' not in request.files:
            return render_template('admin_upload_document.html', error='No file selected')
        
        file = request.files['file']
        if file.filename == '':
            return render_template('admin_upload_document.html', error='No file selected')
        
        if file and allowed_file(file.filename):
            # Create secure filename
            original_filename = file.filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{timestamp}_{original_filename}"
            
            # Save file
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(file_path)
            
            # Get file size
            file_size = os.path.getsize(file_path)
            
            # Save to database
            conn = sqlite3.connect('database.db')
            c = conn.cursor()
            c.execute('''INSERT INTO documents (title, description, filename, file_path, file_size) 
                         VALUES (?, ?, ?, ?, ?)''',
                      (title, description, original_filename, file_path, file_size))
            conn.commit()
            conn.close()
            
            return redirect('/admin/documents')
        else:
            return render_template('admin_upload_document.html', error='Invalid file type. Only PDF files are allowed.')
    
    return render_template('admin_upload_document.html')

@app.route('/admin/documents/download/<int:doc_id>')
def admin_download_document(doc_id):
    if not session.get('admin_logged_in'):
        return redirect('/admin/login')
    
    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    c.execute('SELECT filename, file_path FROM documents WHERE id = ? AND is_active = 1', (doc_id,))
    document = c.fetchone()
    conn.close()
    
    if not document:
        return "Document not found", 404
    
    filename, file_path = document
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True, download_name=filename)
    else:
        return "File not found", 404

@app.route('/admin/documents/delete/<int:doc_id>')
def admin_delete_document(doc_id):
    if not session.get('admin_logged_in'):
        return redirect('/admin/login')
    
    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    
    # Get file path before marking as inactive
    c.execute('SELECT file_path FROM documents WHERE id = ?', (doc_id,))
    result = c.fetchone()
    
    if result:
        file_path = result[0]
        # Mark as inactive in database
        c.execute('UPDATE documents SET is_active = 0 WHERE id = ?', (doc_id,))
        conn.commit()
        
        # Optionally delete the actual file
        # if os.path.exists(file_path):
        #     os.remove(file_path)
    
    conn.close()
    return redirect('/admin/documents')

@app.route('/admin/logout')
def admin_logout():
    session.pop('admin_logged_in', None)
    return redirect('/admin/login')

if __name__ == '__main__':
    init_db()
    app.run(debug=True, host='0.0.0.0')
